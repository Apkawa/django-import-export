from __future__ import unicode_literals
from django.utils.six import moves
import tempfile

import sys
import warnings
import datetime
import tablib

try:
    from tablib.compat import xlrd

    XLS_IMPORT = True
except ImportError:
    try:
        import xlrd  # NOQA

        XLS_IMPORT = True
    except ImportError:
        xls_warning = "Installed `tablib` library does not include"
        "import support for 'xls' format and xlrd module is not found."
        warnings.warn(xls_warning, ImportWarning)
        XLS_IMPORT = False

try:
    import openpyxl

    XLSX_IMPORT = True
except ImportError:
    try:
        from tablib.compat import openpyxl

        XLSX_IMPORT = hasattr(openpyxl, 'load_workbook')
    except ImportError:
        xlsx_warning = "Installed `tablib` library does not include"
        "import support for 'xlsx' format and openpyxl module is not found."
        warnings.warn(xlsx_warning, ImportWarning)
        XLSX_IMPORT = False

try:
    from importlib import import_module
except ImportError:
    from django.utils.importlib import import_module

from django.utils import six


class Format(object):
    def get_title(self):
        return type(self)

    def create_dataset(self, in_stream):
        """
        Create dataset from given string.
        """
        raise NotImplementedError()

    def export_data(self, dataset):
        """
        Returns format representation for given dataset.
        """
        raise NotImplementedError()

    def is_binary(self):
        """
        Returns if this format is binary.
        """
        return True

    def get_read_mode(self):
        """
        Returns mode for opening files.
        """
        return 'rb'

    def get_extension(self):
        """
        Returns extension for this format files.
        """
        return ""

    def get_content_type(self):
        # For content types see
        # http://www.iana.org/assignments/media-types/media-types.xhtml
        return 'application/octet-stream'

    def can_import(self):
        return False

    def can_export(self):
        return False


class TablibFormat(Format):
    TABLIB_MODULE = None
    CONTENT_TYPE = 'application/octet-stream'

    def get_format(self):
        """
        Import and returns tablib module.
        """
        return import_module(self.TABLIB_MODULE)

    def get_title(self):
        return self.get_format().title

    def create_dataset(self, in_stream):
        data = tablib.Dataset()
        self.get_format().import_set(data, in_stream)
        return data

    def export_data(self, dataset):
        return self.get_format().export_set(dataset)

    def get_extension(self):
        # we support both 'extentions' and 'extensions' because currently
        # tablib's master branch uses 'extentions' (which is a typo) but it's
        # dev branch already uses 'extension'.
        # TODO - remove this once the typo is fixxed in tablib's master branch
        if hasattr(self.get_format(), 'extentions'):
            return self.get_format().extentions[0]
        return self.get_format().extensions[0]

    def get_content_type(self):
        return self.CONTENT_TYPE

    def can_import(self):
        return hasattr(self.get_format(), 'import_set')

    def can_export(self):
        return hasattr(self.get_format(), 'export_set')


class TextFormat(TablibFormat):
    def get_read_mode(self):
        return 'rU'

    def is_binary(self):
        return False


class CSV(TextFormat):
    TABLIB_MODULE = 'tablib.formats._csv'
    CONTENT_TYPE = 'text/csv'

    def create_dataset(self, in_stream):
        if sys.version_info[0] < 3:
            # python 2.7 csv does not do unicode
            return super(CSV, self).create_dataset(in_stream.encode('utf-8'))
        return super(CSV, self).create_dataset(in_stream)


class JSON(TextFormat):
    TABLIB_MODULE = 'tablib.formats._json'
    CONTENT_TYPE = 'application/json'


class YAML(TextFormat):
    TABLIB_MODULE = 'tablib.formats._yaml'
    # See http://stackoverflow.com/questions/332129/yaml-mime-type
    CONTENT_TYPE = 'text/yaml'


class TSV(TextFormat):
    TABLIB_MODULE = 'tablib.formats._tsv'
    CONTENT_TYPE = 'text/tab-separated-values'


class ODS(TextFormat):
    TABLIB_MODULE = 'tablib.formats._ods'
    CONTENT_TYPE = 'application/vnd.oasis.opendocument.spreadsheet'


class HTML(TextFormat):
    TABLIB_MODULE = 'tablib.formats._html'
    CONTENT_TYPE = 'text/html'


class XLS(TablibFormat):
    TABLIB_MODULE = 'tablib.formats._xls'
    CONTENT_TYPE = 'application/vnd.ms-excel'

    def can_import(self):
        return XLS_IMPORT

    def create_dataset(self, in_stream):
        """
        Create dataset from first sheet.
        """
        assert XLS_IMPORT

        xls_book = xlrd.open_workbook(file_contents=in_stream)
        dataset = tablib.Dataset()

        sheet = xls_book.sheets()[0]

        dataset.headers = sheet.row_values(0)

        for i in moves.range(1, sheet.nrows):
            row = []
            for c in xrange(sheet.ncols):
                cell = sheet.cell(i, c)
                cell_value = cell.value

                if cell.ctype == xlrd.XL_CELL_NUMBER and int(cell_value) == cell_value:
                    cell_value = int(cell_value)

                elif cell.ctype == xlrd.XL_CELL_DATE:
                    dt_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                    # Create datetime object from this tuple.
                    cell_value = datetime.datetime(
                            dt_tuple[0], dt_tuple[1], dt_tuple[2],
                            dt_tuple[3], dt_tuple[4], dt_tuple[5]
                    )
                row.append(cell_value)
            dataset.append(row)
        return dataset


class XLSX(TablibFormat):
    TABLIB_MODULE = 'tablib.formats._xlsx'
    CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def can_import(self):
        return XLSX_IMPORT

    def create_dataset(self, in_stream):
        """
        Create dataset from first sheet.
        """
        assert XLSX_IMPORT
        from io import BytesIO
        xlsx_book = openpyxl.load_workbook(BytesIO(in_stream))

        dataset = tablib.Dataset()
        sheet = xlsx_book.active

        dataset.headers = [cell.value for cell in sheet.rows[0]]

        for i in moves.range(1, len(sheet.rows)):
            row = []
            for c in xrange(sheet.get_highest_column()):
                cell = sheet.cell(row=i, column=c)
                cell_value = cell.value

                if (cell.data_type == cell.TYPE_NUMERIC
                    and isinstance(cell_value, (int, float, basestring))
                    and int(cell_value) == cell_value):
                    cell_value = int(cell_value)
                row.append(cell_value)
            dataset.append(row)
        return dataset
